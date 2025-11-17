"""
端到端流程：从 Wind 拉取数据，写入 Excel，截取指定区域并发送邮件。
所有步骤都通过明确配置暴露出来，调用方可以控制抓取内容、填入位置、
截图范围以及邮件收件人。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import logging

try:
    from WindPy import w as wind_client  # type: ignore
except ImportError:  # pragma: no cover - WindPy not available outside Wind env
    wind_client = None  # type: ignore

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - only needed when running workflow
    load_workbook = None  # type: ignore

try:
    import win32com.client  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    win32com = None  # type: ignore

try:
    import smtplib
    from email.message import EmailMessage
except ImportError:  # pragma: no cover - extremely unlikely
    smtplib = None  # type: ignore
    EmailMessage = None  # type: ignore

logging.basicConfig(level=logging.INFO)


# --------------------------------------------------------------------------- #
# Configuration models
# --------------------------------------------------------------------------- #


@dataclass
class DataSpec:
    """
    Wind 请求配置：标的、字段、区间、频率等全部集中在这里。

    示例:
        DataSpec(
            identifier="000001.SZ",
            fields=["close", "open"],
            start="2023-01-01",
            end="2023-12-31",
            options="PriceAdj=B"
        )
    """

    identifier: str
    fields: List[str]
    start: str
    end: str
    frequency: str = "D"
    options: Optional[str] = None


@dataclass
class CellMapping:
    """
    Excel 填充配置：告诉程序“哪个字段写到哪个工作簿/工作表/单元格”。
    """

    workbook_path: Path
    sheet_name: str
    mapping: Dict[str, str]


@dataclass
class ScreenshotSpec:
    """
    Excel 截图配置：指定工作簿、工作表以及需要截图的 Range。
    """

    workbook_path: Path
    sheet_name: str
    range_address: str
    output_path: Path


@dataclass
class EmailSpec:
    """
    邮件发送配置：SMTP、账号密码、收件人、正文等。
    """

    smtp_host: str
    smtp_port: int
    username: str
    password: str
    recipients: List[str]
    cc: List[str] = field(default_factory=list)
    bcc: List[str] = field(default_factory=list)
    subject: str = "AutoScraper Report"
    body: str = "Please find the requested data attached."
    attachments: List[Path] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Wind data access
# --------------------------------------------------------------------------- #


class WindDataFetcher:
    """
    WindPy 的轻量封装：负责登录、拉取数据、退出。
    """

    def __init__(self) -> None:
        if wind_client is None:
            raise RuntimeError(
                "WindPy is not available. Install WindPy and run this on a machine "
                "with Wind Financial Terminal."
            )

    def __enter__(self) -> "WindDataFetcher":
        logging.info("Logging into Wind...")
        wind_client.start()
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        logging.info("Logging out of Wind.")
        wind_client.stop()

    def fetch(self, spec: DataSpec) -> Dict[str, List]:
        """
        Fetches data according to the supplied spec and returns a dictionary
        mapping field -> list of values.
        """
        logging.info(
            "Fetching Wind data: %s %s %s-%s",
            spec.identifier,
            spec.fields,
            spec.start,
            spec.end,
        )
        response = wind_client.wsd(
            spec.identifier,
            ",".join(spec.fields),
            spec.start,
            spec.end,
            f"Period={spec.frequency};{spec.options or ''}",
        )
        if response.ErrorCode != 0:
            raise RuntimeError(f"Wind request failed: {response.ErrorCode}")

        payload: Dict[str, List] = {}
        for idx, field in enumerate(spec.fields):
            payload[field] = response.Data[idx]

        # include timestamps when present
        if response.Times:
            payload["times"] = response.Times
        return payload


# --------------------------------------------------------------------------- #
# Excel helpers
# --------------------------------------------------------------------------- #


class ExcelPopulator:
    """
    使用 openpyxl 将数据写回 Excel，避免依赖 Excel 进程。
    """

    def __init__(self) -> None:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for Excel writing.")

    def populate(self, spec: CellMapping, payload: Dict[str, List]) -> Path:
        workbook_path = spec.workbook_path
        logging.info("Populating Excel workbook: %s", workbook_path)
        wb = load_workbook(workbook_path)
        sheet = wb[spec.sheet_name]
        for key, cell in spec.mapping.items():
            if key not in payload:
                logging.warning("Payload does not include key '%s'", key)
                continue
            sheet[cell] = payload[key]
        wb.save(workbook_path)
        return workbook_path


class ExcelScreenshotter:
    """
    借助 Excel COM 截图某个区域，适用于安装了 Excel 的 Windows 环境。
    """

    def __init__(self) -> None:
        if win32com is None:
            raise RuntimeError("pywin32 is required for Excel screenshots.")

    def capture(self, spec: ScreenshotSpec) -> Path:
        workbook_path = Path(spec.workbook_path).resolve()
        output_path = Path(spec.output_path).resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        logging.info(
            "Capturing Excel range %s:%s from %s",
            spec.sheet_name,
            spec.range_address,
            workbook_path,
        )
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        try:
            workbook = excel.Workbooks.Open(str(workbook_path))
            worksheet = workbook.Worksheets(spec.sheet_name)
            worksheet.Range(spec.range_address).CopyPicture()
            chart = workbook.Worksheets(spec.sheet_name).ChartObjects().Add(0, 0, 500, 300)
            chart.Chart.Paste()
            if not chart.Chart.Export(str(output_path)):
                raise RuntimeError(f"Excel 截图导出失败: {output_path}")
            chart.Delete()
            workbook.Close(SaveChanges=0)
        finally:
            excel.Quit()
        return output_path


# --------------------------------------------------------------------------- #
# Email helper
# --------------------------------------------------------------------------- #


class EmailSender:
    """
    SMTP 邮件发送器，支持附带 Excel/图片等附件。
    """

    def __init__(self, spec: EmailSpec) -> None:
        if smtplib is None or EmailMessage is None:
            raise RuntimeError("smtplib/email are required.")
        self.spec = spec

    def send(self) -> None:
        logging.info("Sending email to %s", self.spec.recipients)
        message = EmailMessage()
        message["Subject"] = self.spec.subject
        message["From"] = self.spec.username
        message["To"] = ", ".join(self.spec.recipients)
        if self.spec.cc:
            message["Cc"] = ", ".join(self.spec.cc)
        if self.spec.bcc:
            message["Bcc"] = ", ".join(self.spec.bcc)
        message.set_content(self.spec.body)

        for attachment in self.spec.attachments:
            data = attachment.read_bytes()
            message.add_attachment(
                data,
                maintype="application",
                subtype="octet-stream",
                filename=attachment.name,
            )

        with smtplib.SMTP(self.spec.smtp_host, self.spec.smtp_port) as client:
            client.starttls()
            client.login(self.spec.username, self.spec.password)
            client.send_message(message)


# --------------------------------------------------------------------------- #
# Workflow driver
# --------------------------------------------------------------------------- #


class Workflow:
    """
    流程编排器：串联 Wind 拉数、Excel 填写、截图与邮件。
    """

    def __init__(
        self,
        data_fetcher: WindDataFetcher,
        populator: ExcelPopulator,
        screenshotter: ExcelScreenshotter,
    ) -> None:
        self.data_fetcher = data_fetcher
        self.populator = populator
        self.screenshotter = screenshotter

    def run(
        self,
        data_spec: DataSpec,
        cell_mapping: CellMapping,
        screenshot_spec: ScreenshotSpec,
        email_spec: EmailSpec,
    ) -> None:
        with self.data_fetcher as fetcher:
            payload = fetcher.fetch(data_spec)

        workbook_path = self.populator.populate(cell_mapping, payload)
        screenshot_path = self.screenshotter.capture(screenshot_spec)

        email_spec.attachments.extend([workbook_path, screenshot_path])
        EmailSender(email_spec).send()


# --------------------------------------------------------------------------- #
# Example usage
# --------------------------------------------------------------------------- #


def run_workflow_example() -> None:
    """
    示例入口：演示如何填写配置并跑通整个流程。
    """

    data_spec = DataSpec(
        identifier="000001.SZ",
        fields=["close", "open"],
        start="2024-01-01",
        end="2024-03-31",
    )

    cell_mapping = CellMapping(
        workbook_path=Path("report.xlsx"),
        sheet_name="Sheet1",
        mapping={"close": "B2", "open": "B3"},
    )

    screenshot_spec = ScreenshotSpec(
        workbook_path=cell_mapping.workbook_path,
        sheet_name="Sheet1",
        range_address="A1:F20",
        output_path=Path("report.png"),
    )

    email_spec = EmailSpec(
        smtp_host="smtp.example.com",
        smtp_port=587,
        username="bot@example.com",
        password="password",
        recipients=["recipient@example.com"],
        subject="Daily Wind Report",
        body="See attached Excel and screenshot.",
    )

    workflow = Workflow(
        data_fetcher=WindDataFetcher(),
        populator=ExcelPopulator(),
        screenshotter=ExcelScreenshotter(),
    )
    workflow.run(data_spec, cell_mapping, screenshot_spec, email_spec)


if __name__ == "__main__":
    run_workflow_example()
